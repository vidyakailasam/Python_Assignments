import openpyxl as xl
wb = xl.load_workbook("/home/vidya/PycharmProjects/pythonProject/hotel_info-92.xlsx")
sheet = wb.active
R=sheet.max_row
dup=[]
uni=[]
for i in range (2,R+1):
 x=sheet.cell(row=i,column=1)
 if x.value in uni:
    dup.append(x.value)
 else:
    uni.append(x.value)
print(len(dup))